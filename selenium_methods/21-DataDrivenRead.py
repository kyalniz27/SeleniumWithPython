import openpyxl


path = "C://Users//seker//eclipse-workspace//SeleniumPractice//excel//data.xlsx"
workbook = openpyxl.load_workbook(path)
# 1. Get the sheet from Excel file
sheet = workbook.active
# 2. Second way of getting the sheet from Excel
#sheet = workbook['Sheet1']

rows = sheet.max_row
cols = sheet.max_column

print("Number of rows:",rows)
print("Number of cols:",cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="     ")
    print()

